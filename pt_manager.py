from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from datetime import datetime
import tkinter as tk
from tkinter import messagebox
import os
import sys

if getattr(sys, "frozen", False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

EXCEL_FILE = os.path.join(BASE_DIR, "pt_회원관리.xlsx")

DEFAULT_FONT = Font(name="돋움", size=16)


def get_number(value):
    if value is None:
        return 0
    try:
        return int(value)
    except:
        return 0


def apply_font(ws, row_num):
    for cell in ws[row_num]:
        cell.font = DEFAULT_FONT


def get_sheet(wb, name1, name2):
    if name1 in wb.sheetnames:
        return wb[name1]
    if name2 in wb.sheetnames:
        return wb[name2]
    raise Exception(f"시트를 찾을 수 없습니다: {name1}")


def create_excel_file():
    if os.path.exists(EXCEL_FILE):
        return

    wb = Workbook()

    ws = wb.active
    ws.title = "회원 목록"
    ws.append(["이름", "등록 횟수", "PT사용 횟수", "PT남은 횟수", "금액"])
    apply_font(ws, 1)

    ws2 = wb.create_sheet("PT 수업일")
    ws2.append(["날짜", "시간", "이름", "수업내용"])
    apply_font(ws2, 1)

    ws3 = wb.create_sheet("PT 결제일")
    ws3.append(["결제일", "이름", "등록 횟수", "금액"])
    apply_font(ws3, 1)

    wb.save(EXCEL_FILE)


def add_payment():
    try:
        name = entry_name.get().strip()
        count = int(entry_count.get())
        amount = int(entry_amount.get())

        if not name:
            messagebox.showerror("오류", "이름을 입력하세요.")
            return

        wb = load_workbook(EXCEL_FILE)

        ws_members = get_sheet(wb, "회원 목록", "회원목록")
        ws_payments = get_sheet(wb, "PT 결제일", "PT결제일")

        today = datetime.now().strftime("%Y-%m-%d")

        ws_payments.append([today, name, count, amount])
        apply_font(ws_payments, ws_payments.max_row)

        member_found = False

        for row_num in range(2, ws_members.max_row + 1):
            member_name = ws_members.cell(row=row_num, column=1).value

            if member_name is None:
                continue

            if str(member_name).strip() == name:
                old_total = get_number(ws_members.cell(row=row_num, column=2).value)
                old_remain = get_number(ws_members.cell(row=row_num, column=4).value)
                old_amount = get_number(ws_members.cell(row=row_num, column=5).value)

                ws_members.cell(row=row_num, column=2).value = old_total + count
                ws_members.cell(row=row_num, column=4).value = old_remain + count
                ws_members.cell(row=row_num, column=5).value = old_amount + amount

                member_found = True
                break

        if not member_found:
            ws_members.append([name, count, 0, count, amount])
            apply_font(ws_members, ws_members.max_row)

        wb.save(EXCEL_FILE)

        messagebox.showinfo("완료", f"{name} 회원 결제 등록 완료")

        entry_name.delete(0, tk.END)
        entry_count.delete(0, tk.END)
        entry_amount.delete(0, tk.END)

    except PermissionError:
        messagebox.showerror("오류", "엑셀 파일이 열려 있습니다.\n엑셀을 닫고 다시 시도하세요.")

    except ValueError:
        messagebox.showerror("오류", "등록 횟수와 금액은 숫자로 입력하세요.")

    except Exception as e:
        messagebox.showerror("오류", str(e))


def add_lesson():
    try:
        date = entry_date.get().strip()
        time = entry_time.get().strip()
        name = entry_lesson_name.get().strip()
        content = entry_content.get().strip()

        if not name:
            messagebox.showerror("오류", "회원 이름을 입력하세요.")
            return

        wb = load_workbook(EXCEL_FILE)

        ws_members = get_sheet(wb, "회원 목록", "회원목록")
        ws_lessons = get_sheet(wb, "PT 수업일", "PT수업일")

        for row_num in range(2, ws_members.max_row + 1):
            member_name = ws_members.cell(row=row_num, column=1).value

            if member_name is None:
                continue

            if str(member_name).strip() == name:
                used_count = get_number(ws_members.cell(row=row_num, column=3).value)
                remain_count = get_number(ws_members.cell(row=row_num, column=4).value)

                if remain_count <= 0:
                    messagebox.showerror("오류", "남은 PT 횟수가 없습니다.")
                    return

                ws_lessons.append([date, time, name, content])
                apply_font(ws_lessons, ws_lessons.max_row)

                ws_members.cell(row=row_num, column=3).value = used_count + 1
                ws_members.cell(row=row_num, column=4).value = remain_count - 1

                wb.save(EXCEL_FILE)

                messagebox.showinfo("완료", f"{name} 회원 수업 등록 완료")

                entry_date.delete(0, tk.END)
                entry_time.delete(0, tk.END)
                entry_lesson_name.delete(0, tk.END)
                entry_content.delete(0, tk.END)

                return

        messagebox.showerror("오류", "회원을 찾을 수 없습니다.")

    except PermissionError:
        messagebox.showerror("오류", "엑셀 파일이 열려 있습니다.\n엑셀을 닫고 다시 시도하세요.")

    except Exception as e:
        messagebox.showerror("오류", str(e))


create_excel_file()

root = tk.Tk()
root.title("PT 회원관리 프로그램")
root.geometry("320x520")

tk.Label(root, text="[결제 등록]").pack(pady=5)

tk.Label(root, text="이름").pack()
entry_name = tk.Entry(root)
entry_name.pack()

tk.Label(root, text="등록 횟수").pack()
entry_count = tk.Entry(root)
entry_count.pack()

tk.Label(root, text="금액").pack()
entry_amount = tk.Entry(root)
entry_amount.pack()

tk.Button(root, text="결제 등록", command=add_payment).pack(pady=15)

tk.Label(root, text="----------------").pack()

tk.Label(root, text="[PT 수업 등록]").pack(pady=5)

tk.Label(root, text="날짜 예: 2026-05-11").pack()
entry_date = tk.Entry(root)
entry_date.pack()

tk.Label(root, text="시간 예: 19:00").pack()
entry_time = tk.Entry(root)
entry_time.pack()

tk.Label(root, text="회원 이름").pack()
entry_lesson_name = tk.Entry(root)
entry_lesson_name.pack()

tk.Label(root, text="수업내용").pack()
entry_content = tk.Entry(root)
entry_content.pack()

tk.Button(root, text="수업 등록", command=add_lesson).pack(pady=15)

root.mainloop()